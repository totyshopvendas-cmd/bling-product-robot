"""Pricing table loader (CSV or Excel) and lookup.

Expected columns (header names, any order):
  Custo do Catálogo | Preço da Loja | Preço de Venda

CSV fallback (semicolon):
  Custo do Catálogo;Preço da Loja;Preço de Venda
  21,99;50,50;5050

Preço de Venda is the integer the robot types (5050 = R$ 50,50).
"""
from __future__ import annotations

import asyncio
import csv
import io
import logging
import os
import re
import unicodedata
import urllib.request
from pathlib import Path
from typing import Any, Iterable

from db import db
from models import PriceLookupResponse

logger = logging.getLogger(__name__)

_PROJECT_ROOT = Path(__file__).resolve().parent.parent
_DATA_DIR = _PROJECT_ROOT / "data"
DEFAULT_SHEET_ID = "1ABXgFvtAy7kKmpUl73yYs_69qQneLZr1"

_TABLE_NAMES = [
    "tabela_precos.xlsx",
    "tabela_precos.csv",
    "tabela_precos_johndrop.xlsx",
    "tabela_precos_johndrop.csv",
    "tabela_precos_johndrop_1_00_a_1000_00_centavo_a_centavo.xlsx",
    "tabela_precos_johndrop_1_00_a_1000_00_centavo_a_centavo.csv",
    "tabela_precos_johndrop_1_00_a_1000_00_centavo_a_centavoUTF.csv",
]


def _norm(text: str) -> str:
    raw = unicodedata.normalize("NFD", str(text or ""))
    raw = "".join(ch for ch in raw if unicodedata.category(ch) != "Mn")
    return raw.lower().strip()


def _numeric_str(value: Any) -> str:
    s = unicodedata.normalize("NFKC", str(value or ""))
    s = s.replace("R$", "").replace("r$", "")
    return re.sub(r"[^\d,.\-]", "", s)


def _to_cents(value: Any) -> int:
    if value is None or value == "":
        raise ValueError("custo vazio")
    if isinstance(value, bool):
        raise ValueError("custo inválido")
    if isinstance(value, (int, float)):
        return int(round(float(value) * 100))
    s = _numeric_str(value)
    if not s:
        raise ValueError("custo vazio")
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    return int(round(float(s) * 100))


def _store_brl(value: Any) -> str:
    if value is None or value == "":
        return ""
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return f"{float(value):.2f}".replace(".", ",")
    return str(value).strip().replace("R$", "").strip()


def _sale_int(value: Any) -> int:
    if value is None or value == "":
        raise ValueError("preço de venda vazio")
    if isinstance(value, bool):
        raise ValueError("preço de venda inválido")
    if isinstance(value, int):
        return int(value)
    if isinstance(value, float):
        if value == int(value):
            return int(value)
        return int(round(value * 100))
    s = _numeric_str(value)
    if not s:
        raise ValueError("preço de venda vazio")
    if s.isdigit() or (s.startswith("-") and s[1:].isdigit()):
        return int(s)
    if "," in s:
        return int(round(float(s.replace(".", "").replace(",", ".")) * 100))
    if "." in s:
        f = float(s)
        if f == int(f):
            return int(f)
        return int(round(f * 100))
    return int(float(s))


def _pick_columns(header: list) -> tuple[int, int, int]:
    labels = [_norm(c) for c in header]
    cost = store = sale = None
    for i, lab in enumerate(labels):
        if cost is None and "custo" in lab:
            cost = i
        elif store is None and "loja" in lab:
            store = i
        elif sale is None and "venda" in lab:
            sale = i
    if cost is not None and store is not None and sale is not None:
        return cost, store, sale
    if len(header) >= 3:
        return 0, 1, 2
    raise ValueError("A planilha precisa das colunas Custo, Preço da Loja e Preço de Venda")


def _is_xlsx(content: bytes, filename: str = "") -> bool:
    if content[:2] == b"PK":
        return True
    name = (filename or "").lower()
    return name.endswith(".xlsx") or name.endswith(".xlsm") or name.endswith(".xls")


def _decode_csv(content: bytes) -> str:
    if content.startswith(b"\xff\xfe") or content.startswith(b"\xfe\xff"):
        return content.decode("utf-16")
    for enc in ("utf-8-sig", "cp1252", "latin-1"):
        try:
            return content.decode(enc)
        except UnicodeDecodeError:
            continue
    return content.decode("utf-8", errors="replace")


def _rows_from_csv(content: bytes) -> list[list]:
    """Excel pt-BR often uses ';' — never pick ',' just because decimals contain commas."""
    text = _decode_csv(content)
    best_rows: list[list] = []
    best_score = -1
    for delim in (";", "\t", ","):
        rows = [list(r) for r in csv.reader(io.StringIO(text), delimiter=delim)]
        exact3 = sum(1 for r in rows[:40] if len(r) == 3)
        ge3 = sum(1 for r in rows[:40] if len(r) >= 3)
        score = exact3 * 20 + ge3
        if score > best_score:
            best_score = score
            best_rows = rows
    return best_rows


def _rows_from_xlsx(content: bytes) -> list[list]:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    try:
        ws = wb.active
        return [list(row) for row in ws.iter_rows(values_only=True)]
    finally:
        wb.close()


def _row_has_value(row: list) -> bool:
    return bool(row) and any(c is not None and str(c).strip() != "" for c in row)


def _iter_raw_rows(content: bytes, filename: str = "") -> Iterable[list]:
    looks_xlsx = _is_xlsx(content, filename)
    if looks_xlsx:
        try:
            for row in _rows_from_xlsx(content):
                if _row_has_value(row):
                    yield row
            return
        except Exception:
            if content[:2] == b"PK":
                raise
    for row in _rows_from_csv(content):
        if _row_has_value(row):
            yield row


def _parse_table(content: bytes, filename: str = "") -> tuple[list[dict], list[str]]:
    """CPU-bound parse used in a thread so FastAPI does not freeze."""
    if not content:
        return [], ["Arquivo vazio"]
    try:
        rows = list(_iter_raw_rows(content, filename))
    except Exception as exc:
        return [], [f"Não foi possível ler o arquivo: {exc}"]
    if not rows:
        return [], ["Arquivo vazio"]

    header_idx = None
    cost_i, store_i, sale_i = 0, 1, 2
    for i, row in enumerate(rows[:20]):
        header = [str(c or "") for c in row]
        labs = [_norm(c) for c in header]
        if any("custo" in x for x in labs) or any("venda" in x for x in labs):
            try:
                cost_i, store_i, sale_i = _pick_columns(header)
                header_idx = i
                break
            except Exception:
                continue
    start = (header_idx + 1) if header_idx is not None else 0

    by_cents: dict[int, dict] = {}
    errors: list[str] = []
    for i, row in enumerate(rows[start:], start=start + 1):
        if max(cost_i, store_i, sale_i) >= len(row):
            if len(errors) < 20:
                errors.append(f"L{i}: colunas insuficientes")
            continue
        try:
            cost_cents = _to_cents(row[cost_i])
            by_cents[cost_cents] = {
                "cost_cents": cost_cents,
                "store_price_brl": _store_brl(row[store_i]),
                "sale_price_int": _sale_int(row[sale_i]),
            }
        except Exception as e:
            if len(errors) < 20:
                errors.append(f"L{i}: {e}")
    return list(by_cents.values()), errors


def _is_forbidden_path(path: Path | str) -> bool:
    s = str(path).lower().replace("/", "\\")
    return "meu drive" in s or "google drive" in s


def _user_search_dirs() -> list[Path]:
    home = Path.home()
    dirs = [
        home / "Desktop",
        home / "Downloads",
        home / "Documents",
        home / "OneDrive" / "Desktop",
        home / "OneDrive" / "Downloads",
        home / "TotyShop",
    ]
    out: list[Path] = []
    for d in dirs:
        try:
            if d.is_dir() and not _is_forbidden_path(d):
                out.append(d)
        except Exception:
            continue
    return out


def _candidate_table_paths() -> list[Path]:
    paths: list[Path] = []
    envp = (os.environ.get("PRICING_TABLE_PATH") or "").strip()
    if envp:
        paths.append(Path(envp))
    search = [_DATA_DIR, _PROJECT_ROOT, *_user_search_dirs()]
    for folder in search:
        for name in _TABLE_NAMES:
            paths.append(folder / name)
        try:
            if folder.is_dir():
                paths.extend(sorted(folder.glob("tabela_precos*")))
        except Exception:
            pass
    if _DATA_DIR.is_dir():
        for pattern in ("*.xlsx", "*.csv"):
            paths.extend(sorted(_DATA_DIR.glob(pattern)))
    seen: set[str] = set()
    unique: list[Path] = []
    for p in paths:
        if _is_forbidden_path(p):
            continue
        key = str(p)
        if key in seen:
            continue
        seen.add(key)
        unique.append(p)
    return unique


def _save_local_copy(content: bytes, filename: str = "") -> Path | None:
    try:
        _DATA_DIR.mkdir(parents=True, exist_ok=True)
        dest = _DATA_DIR / ("tabela_precos.xlsx" if _is_xlsx(content, filename) else "tabela_precos.csv")
        dest.write_bytes(content)
        logger.info("Tabela gravada em %s (%s KB)", dest, dest.stat().st_size // 1024)
        return dest
    except Exception as exc:
        logger.warning("não gravou cópia local: %s", exc)
        return None


def _looks_like_html(content: bytes) -> bool:
    head = content[:200].lstrip().lower()
    return head.startswith(b"<!doctype") or head.startswith(b"<html") or b"<title>sign in" in head


def _sheet_id() -> str:
    env = (os.environ.get("PRICING_SHEET_ID") or "").strip()
    if env:
        return env
    url = (os.environ.get("PRICING_SHEET_URL") or "").strip()
    m = re.search(r"/spreadsheets/d/([a-zA-Z0-9-_]+)", url)
    if m:
        return m.group(1)
    return DEFAULT_SHEET_ID


def _download_google() -> tuple[bytes, str]:
    sheet_id = _sheet_id()
    urls = [
        f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx",
        f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv",
        f"https://docs.google.com/spreadsheets/d/{sheet_id}/pub?output=csv",
    ]
    last = "não baixou a planilha"
    for url in urls:
        try:
            req = urllib.request.Request(url, headers={"User-Agent": "TotyShop/1.0"})
            with urllib.request.urlopen(req, timeout=90) as resp:
                data = resp.read()
        except Exception as exc:
            last = str(exc)
            continue
        if not data or len(data) < 40:
            last = "resposta vazia"
            continue
        if _looks_like_html(data):
            last = (
                "A planilha Google não está pública. "
                "Em Compartilhar, escolha Qualquer pessoa com o link."
            )
            continue
        name = "tabela_precos.xlsx" if data[:2] == b"PK" else "tabela_precos.csv"
        logger.info("Tabela baixada do Google (%s, %s KB)", name, len(data) // 1024)
        return data, name
    raise RuntimeError(last)


async def _store_docs(docs: list[dict], errors: list[str]) -> dict:
    if not docs:
        return {
            "imported": 0,
            "errors": errors[:20]
            or ["Nenhuma linha válida. A planilha precisa das colunas Custo, Preço da Loja e Preço de Venda."],
        }
    await db.pricing.delete_many({})
    chunk = 5000
    for i in range(0, len(docs), chunk):
        await db.pricing.insert_many(docs[i : i + chunk], ordered=False)
    return {"imported": len(docs), "errors": errors[:20]}


async def import_table(content: bytes, filename: str = "") -> dict:
    if not content:
        return {"imported": 0, "errors": ["Arquivo vazio"]}
    loop = asyncio.get_running_loop()
    docs, errors = await loop.run_in_executor(None, _parse_table, content, filename)
    res = await _store_docs(docs, errors)
    if res.get("imported"):
        await loop.run_in_executor(None, _save_local_copy, content, filename)
    return res


async def import_csv(content: bytes) -> dict:
    """Backward-compatible alias used by older tests and the upload endpoint."""
    return await import_table(content, filename="pricing.csv")


async def load_bundled_table(force: bool = False) -> dict:
    """Load the JohnDrop price table from disk (no browser upload)."""
    if not force:
        try:
            if await db.pricing.count_documents({}) > 0:
                return {"imported": 0, "skipped": True, "errors": []}
        except Exception as exc:
            logger.warning("pricing count: %s", exc)
    seen: set[str] = set()
    last: dict = {"imported": 0, "errors": ["Nenhuma tabela de preços encontrada"]}
    for path in _candidate_table_paths():
        try:
            resolved = str(path.resolve())
        except Exception:
            resolved = str(path)
        if _is_forbidden_path(resolved) or resolved in seen or not path.is_file():
            continue
        seen.add(resolved)
        try:
            size_kb = path.stat().st_size // 1024
            logger.info("Lendo tabela de preços: %s (%s KB)", resolved, size_kb)
            res = await import_table(path.read_bytes(), filename=path.name)
        except Exception as exc:
            last = {"imported": 0, "errors": [f"{path.name}: {exc}"]}
            continue
        last = {**res, "path": resolved}
        if res.get("imported"):
            logger.info("Tabela de preços: %s linhas de %s", res["imported"], resolved)
            return last
    return last


async def load_now() -> dict:
    """One-click: Google Sheet, then Desktop/Downloads/data. Saves a local copy."""
    errors: list[str] = []
    try:
        loop = asyncio.get_running_loop()
        content, name = await loop.run_in_executor(None, _download_google)
        res = await import_table(content, name)
        if res.get("imported"):
            return {**res, "source": "google"}
        errors.extend(res.get("errors") or [])
    except Exception as exc:
        errors.append(f"Google Planilhas: {exc}")
        logger.warning("download Google da tabela: %s", exc)

    res = await load_bundled_table(force=True)
    if res.get("imported"):
        return {**res, "source": res.get("path") or "arquivo"}
    errors.extend(res.get("errors") or [])
    return {
        "imported": 0,
        "source": "",
        "errors": errors[:20]
        or [
            "Não achei a tabela. Coloque o Excel na pasta data do TotyShop "
            "(não no Google Drive) e clique de novo."
        ],
    }


async def lookup_price(cost: float) -> PriceLookupResponse:
    cost_cents = int(round(cost * 100))
    doc = await db.pricing.find_one({"cost_cents": cost_cents}, {"_id": 0})
    if doc:
        return PriceLookupResponse(
            cost=cost,
            cost_cents=cost_cents,
            sale_price_int=doc["sale_price_int"],
            store_price_brl=doc["store_price_brl"],
            found=True,
        )
    doc = await db.pricing.find_one(
        {"cost_cents": {"$gte": cost_cents}},
        sort=[("cost_cents", 1)],
        projection={"_id": 0},
    )
    if doc:
        return PriceLookupResponse(
            cost=cost,
            cost_cents=cost_cents,
            sale_price_int=doc["sale_price_int"],
            store_price_brl=doc["store_price_brl"],
            found=True,
        )
    return PriceLookupResponse(
        cost=cost,
        cost_cents=cost_cents,
        sale_price_int=0,
        store_price_brl="",
        found=False,
    )


async def stats() -> dict:
    count = await db.pricing.count_documents({})
    sample = await db.pricing.find({}, {"_id": 0}).sort("cost_cents", 1).limit(5).to_list(5)
    last = await db.pricing.find({}, {"_id": 0}).sort("cost_cents", -1).limit(5).to_list(5)
    return {"count": count, "first": sample, "last": last}
